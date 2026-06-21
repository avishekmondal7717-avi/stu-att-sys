# Smart Attendance System — pgvector backend
import psycopg2
import jwt
import secrets
from datetime import datetime, timedelta
from contextlib import asynccontextmanager, contextmanager
from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional

from database import init_db, get_db_connection, release_db_connection, hash_password, verify_password

@contextmanager
def db_session():
    conn = get_db_connection()
    try:
        yield conn
    finally:
        try:
            # Return the connection to the pool instead of closing it directly
            release_db_connection(conn)
        except Exception:
            # Fallback to closing if pool release fails
            try:
                conn.close()
            except Exception:
                pass
import os
import cv2
import numpy as np
import base64
import threading
import time
import math
import hashlib
import smtplib
from email.message import EmailMessage
from pathlib import Path
from fastapi.concurrency import run_in_threadpool
from datetime import timezone, timedelta

# Load environment-based parameters
data_dir = Path(__file__).parent / "data"

det_model = os.environ.get("FACE_DETECTION_MODEL_PATH", "face_detection_yunet_2023mar.onnx")
rec_model = os.environ.get("FACE_RECOGNITION_MODEL_PATH", "face_recognition_sface_2021dec.onnx")

det_model_path = Path(det_model)
if not det_model_path.is_absolute():
    det_model_path = data_dir / det_model

rec_model_path = Path(rec_model)
if not rec_model_path.is_absolute():
    rec_model_path = data_dir / rec_model

score_threshold = float(os.environ.get("FACE_DETECTOR_SCORE_THRESHOLD", "0.9"))
nms_threshold = float(os.environ.get("FACE_DETECTOR_NMS_THRESHOLD", "0.3"))

# Load OpenCV DNN Models once
detector = cv2.FaceDetectorYN_create(
    str(det_model_path), 
    "", (320, 320), score_threshold, nms_threshold, 5000
)
recognizer = cv2.FaceRecognizerSF_create(
    str(rec_model_path), 
    ""
)

# Biometric & Liveness settings
FACE_MATCH_THRESHOLD = float(os.environ.get("FACE_MATCH_THRESHOLD", "0.40"))
SELF_FACE_MATCH_THRESHOLD = float(os.environ.get("SELF_FACE_MATCH_THRESHOLD", "0.48"))
LIVENESS_LAPLACIAN_THRESHOLD = float(os.environ.get("LIVENESS_LAPLACIAN_THRESHOLD", "12.0"))

# Timezone settings (IST default)
LOCAL_OFFSET = float(os.environ.get("LOCAL_TIMEZONE_OFFSET_HOURS", "5.5"))
LOCAL_TZ = timezone(timedelta(hours=LOCAL_OFFSET))

# OpenCV Lock to prevent race conditions on shared detector in thread pool
opencv_lock = threading.Lock()
liveness_lock = threading.Lock()
liveness_challenges = {}
LIVENESS_CHALLENGE_TTL_SECONDS = 30
LIVENESS_YAW_CHANGE = float(os.environ.get("LIVENESS_YAW_CHANGE", "0.10"))

def get_utc_now() -> datetime:
    return datetime.now(timezone.utc)

def format_utc_to_local_time(utc_iso_str: str) -> str:
    if not utc_iso_str or utc_iso_str in ["-", "Pending"]:
        return utc_iso_str
    try:
        dt_utc = datetime.fromisoformat(utc_iso_str.replace("Z", "+00:00"))
        dt_local = dt_utc.astimezone(LOCAL_TZ)
        return dt_local.strftime("%I:%M:%S %p")
    except Exception as e:
        print(f"Error converting time {utc_iso_str}: {e}")
        return utc_iso_str

def format_utc_to_local_date(utc_iso_str: str) -> str:
    if not utc_iso_str or utc_iso_str in ["-", "Pending"]:
        return utc_iso_str
    try:
        dt_utc = datetime.fromisoformat(utc_iso_str.replace("Z", "+00:00"))
        dt_local = dt_utc.astimezone(LOCAL_TZ)
        return dt_local.strftime("%Y-%m-%d")
    except Exception as e:
        print(f"Error converting date {utc_iso_str}: {e}")
        return utc_iso_str

def get_utc_range_for_local_date(local_date_str: str) -> tuple[str, str]:
    local_start = datetime.strptime(local_date_str, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
    local_end = local_start + timedelta(days=1) - timedelta(microseconds=1)
    
    utc_start = local_start.astimezone(timezone.utc)
    utc_end = local_end.astimezone(timezone.utc)
    
    return utc_start.isoformat(), utc_end.isoformat()

def decode_base64_image(b64_string: str) -> np.ndarray:
    if "," in b64_string:
        b64_string = b64_string.split(",")[1]
    img_bytes = base64.b64decode(b64_string)
    nparr = np.frombuffer(img_bytes, np.uint8)
    return cv2.imdecode(nparr, cv2.IMREAD_COLOR)

def extract_face_embedding(frame: np.ndarray):
    with opencv_lock:
        h, w, _ = frame.shape
        detector.setInputSize((w, h))
        _, faces = detector.detect(frame)
        # If initial detection fails, try a fallback: basic enhancement + relaxed detector
        if faces is None or len(faces) == 0:
            try:
                lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
                l, a, b = cv2.split(lab)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                cl = clahe.apply(l)
                lab = cv2.merge((cl, a, b))
                enhanced = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
            except Exception:
                enhanced = frame

            try:
                # Create a temporary detector with a lower score threshold and larger input size
                temp_size = (min(640, max(320, w)), min(640, max(320, h)))
                temp_score = max(0.35, score_threshold * 0.6)
                temp_detector = cv2.FaceDetectorYN_create(
                    str(det_model_path), "", temp_size, temp_score, nms_threshold, 5000
                )
                temp_detector.setInputSize((w, h))
                _, faces = temp_detector.detect(enhanced)
            except Exception:
                faces = None

        if faces is not None and len(faces) > 0:
            face = faces[0]
            box = face[0:4].astype(np.int32)
            landmarks = face[4:14].astype(np.int32)
            x, y, width, height = box
            x1, y1 = max(0, x), max(0, y)
            x2, y2 = min(w, x + width), min(h, y + height)
            face_crop = frame[y1:y2, x1:x2]
            
            is_live = False
            yaw_ratio = 1.0
            if face_crop.size > 0:
                gray_crop = cv2.cvtColor(face_crop, cv2.COLOR_BGR2GRAY)
                laplacian_var = cv2.Laplacian(gray_crop, cv2.CV_64F).var()
                r_eye_x, l_eye_x, nose_x = landmarks[0], landmarks[2], landmarks[4]
                dist_r = abs(nose_x - r_eye_x)
                dist_l = abs(nose_x - l_eye_x)
                yaw_ratio = dist_r / max(dist_l, 0.001)
                is_live = laplacian_var > LIVENESS_LAPLACIAN_THRESHOLD and (0.2 < yaw_ratio < 5.0)

            aligned_face = recognizer.alignCrop(frame, face)
            feature = recognizer.feature(aligned_face)
            return feature.flatten().tolist(), box.tolist(), bool(is_live), {"yaw_ratio": float(yaw_ratio)}
        return None, None, False, {}

def verify_motion_liveness(key: str, yaw_ratio: float, frame_quality_ok: bool):
    """Require one clear left-or-right head movement across frames."""
    now = time.monotonic()
    with liveness_lock:
        state = liveness_challenges.get(key)
        if state and now - state["updated"] > LIVENESS_CHALLENGE_TTL_SECONDS:
            state = None

        if not frame_quality_ok:
            # Head movement naturally creates an occasional blurred frame.
            # Preserve challenge progress instead of forcing the user to start
            # over; the next clear frame must still satisfy the motion test.
            if state:
                state["updated"] = now
                return False, "verifying", "Hold briefly, then turn left or right"
            return False, "verifying", "Use a clearly lit face and hold briefly"

        if not state:
            liveness_challenges[key] = {
                "baseline": yaw_ratio,
                "updated": now,
            }
            return False, "verifying", "Turn your head slightly left or right"

        state["updated"] = now
        baseline = state["baseline"]
        relative_change = (yaw_ratio - baseline) / max(abs(baseline), 0.01)

        if abs(relative_change) < LIVENESS_YAW_CHANGE:
            return False, "verifying", "Turn your head slightly left or right"

        liveness_challenges.pop(key, None)
        return True, "live", "Liveness verified"

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Running migrations/seeding on every dev reload can block the API while the
    # frontend is polling. Run `python database.py` or set RUN_DB_INIT_ON_STARTUP=1
    # when schema initialization is needed.
    if os.environ.get("RUN_DB_INIT_ON_STARTUP") == "1":
        init_db()
    yield

app = FastAPI(title="Smart Attendance System API", lifespan=lifespan)

# Enable CORS for frontend integration
CORS_ORIGINS = [
    origin.strip()
    for origin in os.environ.get(
        "CORS_ORIGINS",
        "http://localhost:5173,http://127.0.0.1:5173",
    ).split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from fastapi.staticfiles import StaticFiles
(data_dir / "uploads").mkdir(parents=True, exist_ok=True)
app.mount("/static/uploads", StaticFiles(directory=str(data_dir / "uploads")), name="uploads")

# JWT Config
JWT_SECRET = os.environ.get("JWT_SECRET")
if not JWT_SECRET or len(JWT_SECRET) < 32:
    raise RuntimeError("JWT_SECRET must be set to a random value of at least 32 characters")
JWT_ALGORITHM = "HS256"

from fastapi.security import OAuth2PasswordBearer

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(hours=24) # 24 hours
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return encoded_jwt

def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    if not token:
        raise HTTPException(
            status_code=401,
            detail="Authentication token is missing",
            headers={"WWW-Authenticate": "Bearer"},
        )
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        email = payload.get("email")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token payload")
            
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
            user = cursor.fetchone()
        
        if not user:
            raise HTTPException(status_code=401, detail="User account not found")
            
        user_dict = row_to_dict(user)
        if user_dict["status"] == "Pending Verification":
            raise HTTPException(status_code=403, detail="Email verification is pending")
        if user_dict["status"] == "Suspended":
            raise HTTPException(status_code=403, detail="Account is suspended")
            
        return user_dict
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired", headers={"WWW-Authenticate": "Bearer"})
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token signature", headers={"WWW-Authenticate": "Bearer"})
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Authentication error: {str(e)}", headers={"WWW-Authenticate": "Bearer"})

def get_current_user_optional(token: Optional[str] = Depends(oauth2_scheme)) -> Optional[dict]:
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except Exception:
        return None

def require_role(allowed_roles: List[str]):
    def dependency(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in allowed_roles:
            raise HTTPException(
                status_code=403,
                detail=f"Access forbidden: role '{current_user['role']}' does not have permission"
            )
        return current_user
    return dependency

def normalize_semester(value) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip().lower()
    if not text:
        return None

    roman_map = {
        "i": "1",
        "ii": "2",
        "iii": "3",
        "iv": "4",
        "v": "5",
        "vi": "6",
        "vii": "7",
        "viii": "8",
    }

    for token in text.replace("-", " ").split():
        clean = token.strip("().,")
        if clean.isdigit() and clean[0] in "12345678":
            return clean[0]
        if clean in roman_map:
            return roman_map[clean]

    return text[0] if text[0] in "12345678" else None

def department_prefix(department: Optional[str]) -> Optional[str]:
    if not department:
        return None

    normalized = str(department).strip().lower()
    if normalized in {"computer science", "computer science & engineering", "computer science and engineering", "cse", "cs"}:
        return "CS"
    if normalized in {"information technology", "it"}:
        return "IT"
    if normalized in {"electronics", "electronics & communication", "electronics and communication", "electronics engineering", "ece", "ec"}:
        return "EC"
    if normalized in {"mechanical engineering", "mechanical", "me"}:
        return "ME"
    if normalized in {"civil engineering", "civil", "ce"}:
        return "CE"
    if normalized in {"electrical engineering", "electrical", "ee"}:
        return "EE"
    if normalized in {"humanities", "hu"}:
        return "HU"

    return None

def log_action(action: str, actor: str, status: str):
    try:
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO audit_logs (action, actor, status) VALUES (%s, %s, %s)",
                (action, actor, status)
            )
            conn.commit()
    except Exception as e:
        print(f"Failed to log action: {e}")

# Models
class StudentCreate(BaseModel):
    rollNumber: str
    fullName: str
    email: Optional[str] = ""
    contact: Optional[str] = ""
    department: str
    course: Optional[str] = "B.Tech"
    semester: Optional[str] = "1"
    gender: Optional[str] = "Male"
    dob: Optional[str] = ""
    address: Optional[str] = ""
    status: Optional[str] = "Active"
    flushBiometrics: Optional[bool] = False


class ScanRequest(BaseModel):
    image: str # Base64 image string
    classCode: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    locationAccuracy: Optional[float] = None

class ToggleSessionRequest(BaseModel):
    classCode: str
    active: bool
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    allowedRadiusMeters: Optional[float] = 100
    locationRequired: Optional[bool] = True


def haversine_distance_meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Return great-circle distance between two WGS84 coordinates."""
    earth_radius = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    return earth_radius * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

class LoginRequest(BaseModel):
    email: str
    password: str
    role: Optional[str] = None

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    password: str

class StudentRegisterRequest(BaseModel):
    rollNumber: str
    fullName: str
    email: str
    contact: Optional[str] = ""
    department: str
    course: Optional[str] = "B.Tech"
    semester: Optional[str] = "1"
    gender: Optional[str] = "Male"
    dob: Optional[str] = ""
    address: Optional[str] = ""
    password: str
    photo: Optional[str] = ""  # base64 face scan from webcam

class TeacherRegisterRequest(BaseModel):
    teacherId: str
    fullName: str
    email: str
    contact: Optional[str] = ""
    department: str
    gender: Optional[str] = "Male"
    dob: Optional[str] = ""
    password: str
    subjects: Optional[List[str]] = None

class TeacherCreate(BaseModel):
    teacherId: str
    fullName: str
    email: str
    contact: Optional[str] = ""
    department: str
    status: Optional[str] = "Active"

# Helper: Convert DB row dict to camelCase, stripping internal-only columns
def row_to_dict(row):
    if not row: return None
    d = dict(row)
    # Never leak raw embedding vectors to API responses
    d.pop('embedding', None)
    mapping = {
        'rollnumber': 'rollNumber',
        'fullname': 'fullName',
        'teacherid': 'teacherId',
        'referenceid': 'referenceId',
        'createdat': 'createdAt',
        'timein': 'timeIn',
        'timeout': 'timeOut',
        'markedby': 'markedBy'
    }
    return {mapping.get(k, k): v for k, v in d.items()}

# ─── Auth APIs ────────────────────────────────────────────────
@app.post("/api/auth/register/student")
async def register_student(payload: StudentRegisterRequest):
    with db_session() as conn:
        cursor = conn.cursor()
        try:
            print(f"[Register] Attempting to register student: {payload.rollNumber}, email: {payload.email}")
            # Check if student already exists
            cursor.execute("SELECT id FROM students WHERE rollNumber = %s OR email = %s", (payload.rollNumber, payload.email))
            if cursor.fetchone():
                print("[Register] Error: Student with this Roll Number or Email already exists")
                raise HTTPException(status_code=400, detail="Student with this Roll Number or Email already exists")
                
            cursor.execute("SELECT id FROM users WHERE email = %s", (payload.email,))
            if cursor.fetchone():
                print("[Register] Error: User email already registered")
                raise HTTPException(status_code=400, detail="User email already registered")

            if not payload.photo:
                print("[Register] Error: Profile photo is missing")
                raise HTTPException(status_code=400, detail="Profile photo (face scan) is mandatory for registration.")
                
            frame = decode_base64_image(payload.photo)
            if frame is None:
                print("[Register] Error: Invalid photo data (failed base64 decoding)")
                raise HTTPException(status_code=400, detail="Invalid photo data.")
                
            embedding, _, _, _ = await run_in_threadpool(extract_face_embedding, frame)
            if not embedding:
                print("[Register] Error: Could not detect a clear face in the photo")
                raise HTTPException(status_code=400, detail="Could not detect a clear face in the provided scan. Please try again.")

            # Upload photo to Neon-friendly S3 / local fallback
            from storage import upload_profile_photo
            photo_url = upload_profile_photo(payload.photo, payload.rollNumber)

            # Insert student record with pgvector embedding
            cursor.execute("""
                INSERT INTO students (rollNumber, fullName, email, contact, department, course, semester, gender, dob, address, photo, status, embedding)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::vector)
                RETURNING id
            """, (
                payload.rollNumber, payload.fullName, payload.email, payload.contact,
                payload.department, payload.course, payload.semester, payload.gender, payload.dob,
                payload.address, photo_url, "Active", str(embedding)
            ))
            
            # Get student DB ID
            student_db_id = cursor.fetchone()[0]
            
            # Insert user login account
            hashed = hash_password(payload.password)
            cursor.execute("""
                INSERT INTO users (email, password, role, referenceId, fullName, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                payload.email, hashed, 'student', payload.rollNumber, payload.fullName, "Active"
            ))
            
            conn.commit()
            print("[Register] Success!")
            return {"success": True, "id": student_db_id, "rollNumber": payload.rollNumber}
        except psycopg2.IntegrityError as ie:
            print(f"[Register] IntegrityError: {ie}")
            raise HTTPException(status_code=400, detail="Database integrity error. Student ID or email might be duplicated.")
        except HTTPException:
            raise
        except Exception as e:
            print(f"[Register] Unexpected exception: {e}")
            raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/register/teacher")
async def register_teacher(payload: TeacherRegisterRequest):
    with db_session() as conn:
        cursor = conn.cursor()
        try:
            # Check if teacher already exists
            cursor.execute("SELECT id FROM teachers WHERE teacherId = %s OR email = %s", (payload.teacherId, payload.email))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="Teacher with this ID or Email already exists")
                
            cursor.execute("SELECT id FROM users WHERE email = %s", (payload.email,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="User email already registered")

            # Insert teacher record (store subjects as comma-separated list)
            subjects_csv = None
            if getattr(payload, 'subjects', None):
                subjects_csv = ",".join([s.strip() for s in payload.subjects if s])

            cursor.execute("""
                INSERT INTO teachers (teacherId, fullName, email, contact, department, subjects, status, photo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                payload.teacherId, payload.fullName, payload.email, payload.contact,
                payload.department, subjects_csv, "Active", f"https://i.pravatar.cc/40?img={hash(payload.teacherId) % 70 + 50}"
            ))
            
            teacher_db_id = cursor.fetchone()[0]
            
            # Insert user login account
            hashed = hash_password(payload.password)
            cursor.execute("""
                INSERT INTO users (email, password, role, referenceId, fullName, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                payload.email, hashed, 'teacher', payload.teacherId, payload.fullName, "Active"
            ))
            
            conn.commit()
            return {"success": True, "id": teacher_db_id, "teacherId": payload.teacherId}
        except psycopg2.IntegrityError:
            raise HTTPException(status_code=400, detail="Database integrity error. Teacher ID or email might be duplicated.")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/login")
async def login(payload: LoginRequest):
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE email = %s", (payload.email,))
        user = cursor.fetchone()
        
        if not user:
            raise HTTPException(status_code=401, detail="Invalid email or password")
            
        user_dict = row_to_dict(user)
        
        if not verify_password(payload.password, user_dict["password"]):
            raise HTTPException(status_code=401, detail="Invalid email or password")
            
        if user_dict["status"] == "Pending Verification":
            raise HTTPException(status_code=400, detail="Email verification pending. Please verify your email first.")
            
        if user_dict["status"] == "Suspended":
            raise HTTPException(status_code=403, detail="Your account has been suspended. Please contact administrator.")
            
        # Check if specific role requested
        if payload.role and payload.role != user_dict["role"]:
            raise HTTPException(status_code=403, detail=f"Unauthorized role. Access denied for role {payload.role}")
            
        # Generate token
        token = create_access_token({
            "email": user_dict["email"],
            "role": user_dict["role"],
            "fullName": user_dict["fullName"],
            "referenceId": user_dict["referenceId"]
        })
        
        log_action("Logged into system", user_dict["fullName"], "Success")
        
        # Fetch extra profile details based on role
        profile_data = {}
        if user_dict["role"] == "student":
            cursor.execute('SELECT * FROM students WHERE email = %s', (user_dict["email"],))
            student_row = cursor.fetchone()
            if student_row:
                student_dict = row_to_dict(student_row)
                profile_data = {
                    "id": student_dict.get("id"),
                    "rollNumber": student_dict.get("rollNumber"),
                    "contact": student_dict.get("contact"),
                    "department": student_dict.get("department"),
                    "course": student_dict.get("course"),
                    "semester": student_dict.get("semester"),
                    "gender": student_dict.get("gender"),
                    "dob": student_dict.get("dob"),
                    "address": student_dict.get("address"),
                    "photo": student_dict.get("photo")
                }
        elif user_dict["role"] == "teacher":
            cursor.execute('SELECT * FROM teachers WHERE email = %s', (user_dict["email"],))
            teacher_row = cursor.fetchone()
            if teacher_row:
                teacher_dict = row_to_dict(teacher_row)
                profile_data = {
                    "teacherId": teacher_dict.get("teacherId"),
                    "contact": teacher_dict.get("contact"),
                    "department": teacher_dict.get("department"),
                    "gender": teacher_dict.get("gender"),
                    "dob": teacher_dict.get("dob")
                }
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "email": user_dict["email"],
                "role": user_dict["role"],
                "fullName": user_dict["fullName"],
                "referenceId": user_dict["referenceId"],
                **profile_data
            }
        }


def send_password_reset_email(recipient: str, reset_url: str):
    host = os.environ.get("SMTP_HOST")
    if not host:
        print(f"[PasswordReset] Development reset link for {recipient}: {reset_url}")
        return False
    message = EmailMessage()
    message["Subject"] = "Smart Attendance password reset"
    message["From"] = os.environ.get("SMTP_FROM", os.environ.get("SMTP_USERNAME", "noreply@example.com"))
    message["To"] = recipient
    message.set_content(f"Reset your password within 15 minutes:\n\n{reset_url}\n\nIgnore this message if you did not request it.")
    port = int(os.environ.get("SMTP_PORT", "587"))
    with smtplib.SMTP(host, port, timeout=10) as smtp:
        smtp.starttls()
        username = os.environ.get("SMTP_USERNAME")
        password = os.environ.get("SMTP_PASSWORD")
        if username and password:
            smtp.login(username, password)
        smtp.send_message(message)
    return True


@app.post("/api/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordRequest):
    neutral = {"message": "If the account exists, password reset instructions have been generated."}
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT email FROM users WHERE LOWER(email) = LOWER(%s)", (payload.email.strip(),))
        user = cursor.fetchone()
        if not user:
            return neutral
        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
        cursor.execute("UPDATE password_reset_tokens SET used_at = CURRENT_TIMESTAMP WHERE user_email = %s AND used_at IS NULL", (user["email"],))
        cursor.execute("""
            INSERT INTO password_reset_tokens (user_email, token_hash, expires_at)
            VALUES (%s, %s, CURRENT_TIMESTAMP + INTERVAL '15 minutes')
        """, (user["email"], token_hash))
        conn.commit()
    reset_url = f"{os.environ.get('FRONTEND_URL', 'http://localhost:5173')}/reset-password?token={raw_token}"
    try:
        sent = await run_in_threadpool(send_password_reset_email, user["email"], reset_url)
    except Exception as exc:
        print(f"[PasswordReset] Email delivery failed: {exc}; reset link: {reset_url}")
        sent = False
    response = {**neutral, "delivery": "email" if sent else "development"}
    if os.environ.get("RESET_DEBUG", "1") == "1":
        response["debugResetToken"] = raw_token
    return response


@app.post("/api/auth/reset-password")
async def reset_password(payload: ResetPasswordRequest):
    if len(payload.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters long.")
    token_hash = hashlib.sha256(payload.token.encode()).hexdigest()
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, user_email FROM password_reset_tokens
            WHERE token_hash = %s AND used_at IS NULL AND expires_at > CURRENT_TIMESTAMP
            FOR UPDATE
        """, (token_hash,))
        reset = cursor.fetchone()
        if not reset:
            raise HTTPException(status_code=400, detail="This reset link is invalid or has expired.")
        cursor.execute("UPDATE users SET password = %s WHERE email = %s", (hash_password(payload.password), reset["user_email"]))
        cursor.execute("UPDATE password_reset_tokens SET used_at = CURRENT_TIMESTAMP WHERE id = %s", (reset["id"],))
        conn.commit()
    log_action("Password reset completed", reset["user_email"], "Success")
    return {"success": True}

@app.get("/api/admin/audit-logs", dependencies=[Depends(require_role(["admin"]))])
async def get_audit_logs():
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, timestamp, action, actor, status FROM audit_logs ORDER BY id DESC LIMIT 50")
        rows = cursor.fetchall()
    
    logs = []
    for r in rows:
        logs.append({
            "id": r["id"],
            "timestamp": r["timestamp"].strftime("%Y-%m-%d %I:%M:%S %p") if r["timestamp"] else "",
            "action": r["action"],
            "actor": r["actor"],
            "status": r["status"]
        })
    
    if not logs:
        logs = [
            {
                "id": 1,
                "timestamp": datetime.now().strftime("%Y-%m-%d %I:%M:%S %p"),
                "action": "System Database Initialized",
                "actor": "System",
                "status": "Success"
            }
        ]
    return {"logs": logs}

# ─── Dashboard Stats API ──────────────────────────────────────
@app.get("/api/dashboard/stats", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def get_dashboard_stats():
    with db_session() as conn:
        cursor = conn.cursor()
        
        # Consistent Timezone Handling: Determine today's date in local time
        today_local = datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")
        utc_start, utc_end = get_utc_range_for_local_date(today_local)
        
        # 1. Total Students
        cursor.execute("SELECT COUNT(*) FROM students")
        total_students = cursor.fetchone()[0]
        
        # 2. Total Present today
        cursor.execute("""
            SELECT COUNT(*) FROM attendance 
            WHERE (date = %s OR (timein >= %s AND timein <= %s)) AND status = 'Present'
        """, (today_local, utc_start, utc_end))
        total_present = cursor.fetchone()[0]
        
        # 3. Total Absent today (Active students - Present today)
        cursor.execute("SELECT COUNT(*) FROM students WHERE status = 'Active'")
        active_students = cursor.fetchone()[0]
        total_absent = max(0, active_students - total_present)
        
        # 4. Attendance Rate
        attendance_rate = int((total_present / max(1, active_students)) * 100)
        
        # 5. Recent Attendance Logs
        cursor.execute("""
            SELECT a.id, a.rollNumber, a.studentName, a.department, a.date, a.timeIn, a.timeOut, a.status, a.markedBy
            FROM attendance a
            ORDER BY a.id DESC LIMIT 10
        """)
        recent_raw = [row_to_dict(r) for r in cursor.fetchall()]
        
        # Format UTC timestamps to local timezone strings on presentation
        recent = []
        for r in recent_raw:
            time_in_local = format_utc_to_local_time(r["timeIn"])
            time_out_local = format_utc_to_local_time(r["timeOut"])
            date_local = format_utc_to_local_date(r["timeIn"]) if r["timeIn"] not in ["-", "Pending"] else r["date"]
            recent.append({
                **r,
                "date": date_local,
                "timeIn": time_in_local,
                "timeOut": time_out_local
            })
    
    return {
        "totalStudents": total_students,
        "totalPresent": total_present,
        "totalAbsent": total_absent,
        "attendanceRate": attendance_rate,
        "recentAttendance": recent
    }

# ─── Students CRUD ───────────────────────────────────────────
@app.get("/api/students", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def get_students(
    department: Optional[str] = None,
    semester: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    with db_session() as conn:
        cursor = conn.cursor()

        if current_user.get("role") == "teacher":
            cursor.execute("SELECT department FROM teachers WHERE email = %s", (current_user.get("email"),))
            teacher = cursor.fetchone()
            if not teacher:
                raise HTTPException(status_code=403, detail="Teacher profile not found")
            department = teacher["department"]
        
        query = "SELECT * FROM students WHERE 1=1"
        params = []
        if department:
            query += " AND department = %s"
            params.append(department)
        if semester:
            query += " AND semester = %s"
            params.append(semester)
            
        cursor.execute(query, params)
        students = [row_to_dict(r) for r in cursor.fetchall()]
    return {"data": students, "total": len(students)}

@app.get("/api/students/{student_id}", dependencies=[Depends(require_role(["admin", "teacher", "student"]))])
async def get_student(student_id: int, current_user: dict = Depends(get_current_user)):
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM students WHERE id = %s", (student_id,))
        student = row_to_dict(cursor.fetchone())
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if current_user.get("role") == "student" and student["rollNumber"] != current_user.get("referenceId"):
        raise HTTPException(status_code=403, detail="You can only view your own student profile")
    if current_user.get("role") == "teacher":
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT department FROM teachers WHERE email = %s", (current_user.get("email"),))
            teacher = cursor.fetchone()
        if not teacher or student["department"] != teacher["department"]:
            raise HTTPException(status_code=403, detail="You can only view students in your department")
    return student

@app.post("/api/students", dependencies=[Depends(require_role(["admin"]))])
async def create_student(student: StudentCreate):
    raise HTTPException(status_code=403, detail="Students must create their own account through registration.")

    # Retained below temporarily for database migration history; unreachable by design.
    with db_session() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO students (rollNumber, fullName, email, contact, department, course, semester, gender, dob, address, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                student.rollNumber, student.fullName, student.email, student.contact,
                student.department, student.course, student.semester, student.gender, student.dob,
                student.address, student.status
            ))
            
            # Get student DB ID
            new_id = cursor.fetchone()[0]
            
            # Auto-create user credentials for the newly created student
            temporary_password = secrets.token_urlsafe(12)
            default_pass = hash_password(temporary_password)
            cursor.execute("""
                INSERT INTO users (email, password, role, referenceId, fullName, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                student.email, default_pass, 'student', student.rollNumber, student.fullName, student.status
            ))
            
            conn.commit()
            return {
                "id": new_id,
                "rollNumber": student.rollNumber,
                "fullName": student.fullName,
                "temporaryPassword": temporary_password,
            }
        except psycopg2.IntegrityError:
            raise HTTPException(status_code=400, detail="Student with this Roll Number or Email already exists")

@app.put("/api/students/{student_id}", dependencies=[Depends(require_role(["admin"]))])
async def update_student(student_id: int, student: StudentCreate, current_user: dict = Depends(get_current_user)):
    with db_session() as conn:
        cursor = conn.cursor()
        
        cursor.execute("SELECT email, status FROM students WHERE id = %s", (student_id,))
        old_student = cursor.fetchone()
        if not old_student:
            raise HTTPException(status_code=404, detail="Student not found")
            
        old_email = old_student["email"]
        old_status = old_student["status"]
        
        if student.flushBiometrics:
            cursor.execute("""
                UPDATE students
                SET rollNumber=%s, fullName=%s, email=%s, contact=%s, department=%s, course=%s, semester=%s, gender=%s, dob=%s, address=%s, status=%s, embedding=NULL
                WHERE id=%s
            """, (
                student.rollNumber, student.fullName, student.email, student.contact,
                student.department, student.course, student.semester, student.gender, student.dob,
                student.address, 'Pending Verification', student_id
            ))
            
            cursor.execute("""
                UPDATE users
                SET email=%s, fullName=%s, referenceId=%s, status=%s
                WHERE email=%s
            """, (
                student.email, student.fullName, student.rollNumber, 'Pending Verification', old_email
            ))
            
            log_action(
                f"Flushed biometric template for student {student.fullName} ({student.rollNumber})",
                current_user["fullName"],
                "Success"
            )
        else:
            cursor.execute("""
                UPDATE students
                SET rollNumber=%s, fullName=%s, email=%s, contact=%s, department=%s, course=%s, semester=%s, gender=%s, dob=%s, address=%s, status=%s
                WHERE id=%s
            """, (
                student.rollNumber, student.fullName, student.email, student.contact,
                student.department, student.course, student.semester, student.gender, student.dob,
                student.address, student.status, student_id
            ))
            
            cursor.execute("""
                UPDATE users
                SET email=%s, fullName=%s, referenceId=%s, status=%s
                WHERE email=%s
            """, (
                student.email, student.fullName, student.rollNumber, student.status, old_email
            ))
            
            if old_status != student.status:
                log_action(
                    f"Toggled status for student {student.fullName} ({student.rollNumber}) to {student.status}",
                    current_user["fullName"],
                    "Success"
                )
                
        conn.commit()
    return {"success": True}


@app.delete("/api/students/{student_id}", dependencies=[Depends(require_role(["admin"]))])
async def delete_student(student_id: int):
    with db_session() as conn:
        cursor = conn.cursor()
        
        # Get student details to clean up credentials
        cursor.execute("SELECT email FROM students WHERE id = %s", (student_id,))
        student = cursor.fetchone()
        if student:
            email = student["email"]
            cursor.execute("DELETE FROM students WHERE id = %s", (student_id,))
            cursor.execute("DELETE FROM users WHERE email = %s", (email,))
            conn.commit()
        
    return {"success": True}

# ─── Teachers CRUD (Admin Hub) ───────────────────────────────
@app.get("/api/teachers", dependencies=[Depends(require_role(["admin"]))])
async def get_teachers(department: Optional[str] = None):
    with db_session() as conn:
        cursor = conn.cursor()
        
        query = "SELECT * FROM teachers WHERE 1=1"
        params = []
        if department:
            query += " AND department = %s"
            params.append(department)
            
        cursor.execute(query, params)
        teachers = [row_to_dict(r) for r in cursor.fetchall()]
    return {"data": teachers, "total": len(teachers)}

@app.get("/api/teachers/{teacher_id}", dependencies=[Depends(require_role(["admin"]))])
async def get_teacher(teacher_id: int):
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM teachers WHERE id = %s", (teacher_id,))
        teacher = row_to_dict(cursor.fetchone())
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    return teacher

@app.post("/api/teachers", dependencies=[Depends(require_role(["admin"]))])
async def create_teacher(teacher: TeacherCreate):
    raise HTTPException(status_code=403, detail="Teachers must create their own account through registration.")

    # Retained below temporarily for database migration history; unreachable by design.
    with db_session() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO teachers (teacherId, fullName, email, contact, department, status, photo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                teacher.teacherId, teacher.fullName, teacher.email, teacher.contact,
                teacher.department, teacher.status, f"https://i.pravatar.cc/40?img={hash(teacher.teacherId) % 30 + 50}"
            ))
            
            # Auto-create user credentials for the newly created teacher
            temporary_password = secrets.token_urlsafe(12)
            default_pass = hash_password(temporary_password)
            cursor.execute("""
                INSERT INTO users (email, password, role, referenceId, fullName, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                teacher.email, default_pass, 'teacher', teacher.teacherId, teacher.fullName, teacher.status
            ))
            
            conn.commit()
            new_id = cursor.lastrowid
            return {
                "id": new_id,
                "teacherId": teacher.teacherId,
                "fullName": teacher.fullName,
                "temporaryPassword": temporary_password,
            }
        except psycopg2.IntegrityError:
            raise HTTPException(status_code=400, detail="Teacher with this ID or Email already exists")

@app.put("/api/teachers/{teacher_id}", dependencies=[Depends(require_role(["admin"]))])
async def update_teacher(teacher_id: int, teacher: TeacherCreate):
    with db_session() as conn:
        cursor = conn.cursor()
        
        cursor.execute("SELECT email FROM teachers WHERE id = %s", (teacher_id,))
        old_teacher = cursor.fetchone()
        if not old_teacher:
            raise HTTPException(status_code=404, detail="Teacher not found")
            
        old_email = old_teacher["email"]
        
        cursor.execute("""
            UPDATE teachers
            SET teacherId=%s, fullName=%s, email=%s, contact=%s, department=%s, status=%s
            WHERE id=%s
        """, (
            teacher.teacherId, teacher.fullName, teacher.email, teacher.contact,
            teacher.department, teacher.status, teacher_id
        ))
        
        # Also update user login profile
        cursor.execute("""
            UPDATE users
            SET email=%s, fullName=%s, referenceId=%s, status=%s
            WHERE email=%s
        """, (
            teacher.email, teacher.fullName, teacher.teacherId, teacher.status, old_email
        ))
        
        conn.commit()
    return {"success": True}

@app.delete("/api/teachers/{teacher_id}", dependencies=[Depends(require_role(["admin"]))])
async def delete_teacher(teacher_id: int):
    with db_session() as conn:
        cursor = conn.cursor()
        
        cursor.execute("SELECT email FROM teachers WHERE id = %s", (teacher_id,))
        teacher = cursor.fetchone()
        if teacher:
            email = teacher["email"]
            cursor.execute("DELETE FROM teachers WHERE id = %s", (teacher_id,))
            cursor.execute("DELETE FROM users WHERE email = %s", (email,))
            conn.commit()
        
    return {"success": True}

# ─── Face Biometrics Enrollment (pgvector) ────────────────────
class FaceEnrollRequest(BaseModel):
    images: List[str]  # List of base64 image strings

@app.post("/api/students/{student_id}/face-images", dependencies=[Depends(require_role(["admin", "student"]))])
async def enroll_student_faces(
    student_id: int,
    request: FaceEnrollRequest,
    current_user: dict = Depends(get_current_user),
):
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, rollnumber FROM students WHERE id = %s", (student_id,))
        student = cursor.fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
        if current_user.get("role") == "student" and student["rollnumber"] != current_user.get("referenceId"):
            raise HTTPException(status_code=403, detail="You can only enroll your own face")

        # Average several clear samples for a more stable template across
        # small pose and lighting changes.
        embeddings = []
        for b64_img in request.images:
            frame = decode_base64_image(b64_img)
            if frame is None:
                continue
            embedding, _, _, _ = await run_in_threadpool(extract_face_embedding, frame)
            if embedding:
                embeddings.append(np.asarray(embedding, dtype=np.float32))

        if embeddings:
            averaged = np.mean(embeddings, axis=0)
            norm = np.linalg.norm(averaged)
            if norm > 0:
                averaged = averaged / norm
            cursor.execute(
                "UPDATE students SET embedding = %s::vector, status = 'Active' WHERE id = %s",
                (str(averaged.tolist()), student_id)
            )
            cursor.execute("""
                UPDATE users SET status = 'Active'
                WHERE role = 'student' AND referenceid = %s
            """, (student["rollnumber"],))
            conn.commit()
            return {"success": True, "samples_enrolled": len(embeddings)}

    raise HTTPException(status_code=400, detail="Could not detect any faces. Ensure proper lighting and a clear view.")


@app.post("/api/student/face-images", dependencies=[Depends(require_role(["student"]))])
async def enroll_current_student_faces(
    request: FaceEnrollRequest,
    current_user: dict = Depends(get_current_user),
):
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM students WHERE rollnumber = %s",
            (current_user.get("referenceId"),),
        )
        student = cursor.fetchone()
    if not student:
        raise HTTPException(status_code=404, detail="Student profile not found")
    return await enroll_student_faces(student["id"], request, current_user)

# ─── Attendance Records ───────────────────────────────────────
@app.get("/api/attendance/session-history", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def get_attendance_session_history(
    period: str = "day",
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    anchor = date or datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")
    if period not in {"day", "month", "year"}:
        raise HTTPException(status_code=400, detail="Period must be day, month, or year")

    filters = []
    params = []
    if current_user.get("role") == "teacher":
        filters.append("s.teacher_email = %s")
        params.append(current_user.get("email"))

    if period == "day":
        filters.append("s.session_date = %s")
        params.append(anchor)
    elif period == "month":
        filters.append("LEFT(s.session_date, 7) = %s")
        params.append(anchor[:7])
    else:
        filters.append("LEFT(s.session_date, 4) = %s")
        params.append(anchor[:4])

    where_sql = " AND ".join(filters) if filters else "TRUE"
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute(f"""
            SELECT
                s.id, s.classcode, s.classname, s.department, s.semester,
                s.teacher_name, s.session_date, s.started_at, s.ended_at, s.isactive,
                COUNT(DISTINCT r.rollnumber) AS total_students,
                COUNT(DISTINCT a.rollnumber) FILTER (WHERE a.status = 'Present') AS present_students
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            WHERE {where_sql}
            GROUP BY s.id
            ORDER BY s.session_date DESC, s.started_at DESC
        """, tuple(params))
        rows = cursor.fetchall()

    sessions = []
    for row in rows:
        total = int(row["total_students"] or 0)
        present = int(row["present_students"] or 0)
        started_at = row["started_at"]
        ended_at = row["ended_at"]
        sessions.append({
            "id": row["id"],
            "classCode": row["classcode"],
            "className": row["classname"],
            "department": row["department"],
            "semester": row["semester"],
            "teacherName": row["teacher_name"],
            "date": row["session_date"],
            "startedAt": started_at.astimezone(LOCAL_TZ).isoformat() if started_at else None,
            "endedAt": ended_at.astimezone(LOCAL_TZ).isoformat() if ended_at else None,
            "isActive": row["isactive"],
            "total": total,
            "present": present,
            "absent": max(total - present, 0),
        })
    return {"sessions": sessions}


@app.get("/api/attendance/session-history/{session_id}", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def get_attendance_session_detail(
    session_id: int,
    current_user: dict = Depends(get_current_user),
):
    with db_session() as conn:
        cursor = conn.cursor()
        owner_filter = ""
        params = [session_id]
        if current_user.get("role") == "teacher":
            owner_filter = " AND teacher_email = %s"
            params.append(current_user.get("email"))
        cursor.execute(f"""
            SELECT * FROM attendance_sessions
            WHERE id = %s {owner_filter}
        """, tuple(params))
        session = cursor.fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Attendance session not found")

        cursor.execute("""
            SELECT
                COALESCE(r.rollnumber, a.rollnumber) AS rollnumber,
                COALESCE(r.studentname, a.studentname) AS studentname,
                COALESCE(r.department, a.department) AS department,
                COALESCE(r.semester, student.semester::text) AS semester,
                a.id, a.timein, a.timeout, a.status, a.markedby
            FROM attendance_session_roster r
            FULL OUTER JOIN attendance a
              ON a.sessionid = r.sessionid AND a.rollnumber = r.rollnumber
            LEFT JOIN students student
              ON student.rollnumber = COALESCE(r.rollnumber, a.rollnumber)
            WHERE COALESCE(r.sessionid, a.sessionid) = %s
            ORDER BY COALESCE(r.rollnumber, a.rollnumber)
        """, (session_id,))
        rows = cursor.fetchall()

    records = []
    for row in rows:
        records.append({
            "id": row["id"] or f"absent-{session_id}-{row['rollnumber']}",
            "rollNumber": row["rollnumber"],
            "studentName": row["studentname"],
            "department": row["department"],
            "semester": row["semester"],
            "timeIn": format_utc_to_local_time(row["timein"]),
            "timeOut": format_utc_to_local_time(row["timeout"]),
            "status": row["status"] or "Absent",
            "markedBy": row["markedby"] or "-",
        })

    return {
        "session": {
            "id": session["id"],
            "classCode": session["classcode"],
            "className": session["classname"],
            "department": session["department"],
            "semester": session["semester"],
            "teacherName": session["teacher_name"],
            "date": session["session_date"],
            "isActive": session["isactive"],
        },
        "data": records,
    }

@app.get("/api/student/attendance", dependencies=[Depends(require_role(["student"]))])
async def get_student_attendance(current_user: dict = Depends(get_current_user)):
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT s.id AS sessionid, s.classcode, s.classname, s.teacher_name,
                   s.session_date, s.started_at, s.ended_at, s.isactive,
                   a.id AS attendanceid, a.timein, a.timeout, a.markedby,
                   a.verification_method, a.face_confidence, a.distance_meters,
                   CASE
                       WHEN a.status = 'Present' THEN 'Present'
                       WHEN s.isactive = TRUE THEN 'Pending'
                       ELSE 'Absent'
                   END AS computed_status
            FROM attendance_session_roster roster
            JOIN attendance_sessions s ON s.id = roster.sessionid
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = roster.rollnumber
            WHERE roster.rollnumber = %s
            ORDER BY s.started_at DESC, s.id DESC
        """, (current_user["referenceId"],))
        rows = cursor.fetchall()

    logs = []
    present_count = 0
    absent_count = 0
    pending_count = 0
    for r in rows:
        status = r["computed_status"]
        if status == "Present": present_count += 1
        elif status == "Absent": absent_count += 1
        else: pending_count += 1
        logs.append({
            "key": f"session-{r['sessionid']}",
            "id": r["attendanceid"],
            "sessionId": r["sessionid"],
            "classCode": r["classcode"],
            "className": r["classname"],
            "teacherName": r["teacher_name"],
            "date": r["session_date"],
            "timeIn": format_utc_to_local_time(r["timein"]),
            "timeOut": format_utc_to_local_time(r["timeout"]),
            "status": status,
            "isActive": r["isactive"],
            "type": f"{r['classname']} ({r['classcode']})",
            "markedBy": r["markedby"] or "-",
            "verificationMethod": r["verification_method"],
            "faceConfidence": r["face_confidence"],
            "distanceMeters": r["distance_meters"],
        })
    completed = present_count + absent_count
    return {
        "data": logs,
        "summary": {
            "present": present_count,
            "absent": absent_count,
            "pending": pending_count,
            "completed": completed,
            "attendanceRate": round((present_count / completed * 100), 1) if completed else 0.0,
        },
    }

# ─── Reports and Exports ─────────────────────────────────────
@app.get("/api/reports/stats", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def get_reports_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    department: Optional[str] = None,
    group_by: Optional[str] = "day",
    current_user: dict = Depends(get_current_user),
):
    if not start_date:
        start_date = (datetime.now(LOCAL_TZ) - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")
        
    with db_session() as conn:
        cursor = conn.cursor()
        session_filters = ["s.session_date BETWEEN %s AND %s"]
        session_params = [start_date, end_date]
        if current_user.get("role") == "teacher":
            cursor.execute("SELECT department FROM teachers WHERE email = %s", (current_user.get("email"),))
            teacher = cursor.fetchone()
            teacher_department = teacher["department"] if teacher else ""
            if department and department.lower() != teacher_department.lower():
                raise HTTPException(status_code=403, detail="Teachers can only view their own department analytics.")
            session_filters.append("LOWER(s.department) = LOWER(%s)")
            session_params.append(teacher_department)
        elif department:
            session_filters.append("LOWER(s.department) = LOWER(%s)")
            session_params.append(department)
        where_sql = " AND ".join(session_filters)

        cursor.execute(f"""
            SELECT
                COUNT(DISTINCT r.studentid) AS students,
                COUNT(r.rollnumber) AS roster_entries,
                COUNT(a.rollnumber) FILTER (WHERE a.status = 'Present') AS present
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            WHERE {where_sql}
        """, tuple(session_params))
        totals = cursor.fetchone()
        total_students = int(totals["students"] or 0)
        total_records = int(totals["roster_entries"] or 0)
        total_present = int(totals["present"] or 0)
        total_absent = max(total_records - total_present, 0)
        avg_attendance = round((total_present / total_records * 100), 1) if total_records else 0.0

        cursor.execute(f"""
            SELECT
                s.department,
                COUNT(DISTINCT r.studentid) AS students,
                COUNT(r.rollnumber) AS roster_entries,
                COUNT(a.rollnumber) FILTER (WHERE a.status = 'Present') AS present
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            WHERE {where_sql}
            GROUP BY s.department
            ORDER BY s.department
        """, tuple(session_params))
        dept_stats = []
        for row in cursor.fetchall():
            roster_entries = int(row["roster_entries"] or 0)
            present = int(row["present"] or 0)
            absent = max(roster_entries - present, 0)
            dept_stats.append({
                "name": row["department"],
                "totalStudents": int(row["students"] or 0),
                "present": present,
                "absent": absent,
                "percentage": round((present / roster_entries * 100), 1) if roster_entries else 0.0,
            })

        grouping = {"day": "s.session_date", "month": "SUBSTRING(s.session_date, 1, 7)", "year": "SUBSTRING(s.session_date, 1, 4)"}.get(group_by or "day", "s.session_date")
        cursor.execute(f"""
            SELECT
                {grouping} AS date,
                COUNT(r.rollnumber) AS total,
                COUNT(a.rollnumber) FILTER (WHERE a.status = 'Present') AS present
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            WHERE {where_sql}
            GROUP BY {grouping}
            ORDER BY {grouping}
        """, tuple(session_params))
        daily_rows = cursor.fetchall()
        
        trend = []
        for r in daily_rows:
            pct = round((r["present"] / r["total"] * 100), 1) if r["total"] else 0.0
            trend.append({
                "date": (datetime.strptime(r["date"], "%Y-%m-%d").strftime("%d %b") if r["date"] and len(r["date"]) == 10 else datetime.strptime(r["date"], "%Y-%m").strftime("%b %Y") if r["date"] and len(r["date"]) == 7 else r["date"] or "-"),
                "value": pct
            })
            
    return {
        "totalStudents": total_students,
        "avgAttendance": f"{avg_attendance}%",
        "presentCount": total_present,
        "absentCount": total_absent,
        "departmentStats": dept_stats,
        "attendanceOverview": trend,
        "period": {"startDate": start_date, "endDate": end_date, "groupBy": group_by},
    }

@app.get("/api/reports/departments", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def get_report_departments(current_user: dict = Depends(get_current_user)):
    with db_session() as conn:
        cursor = conn.cursor()
        if current_user.get("role") == "teacher":
            cursor.execute("SELECT department FROM teachers WHERE email = %s", (current_user.get("email"),))
        else:
            cursor.execute("SELECT DISTINCT department FROM (SELECT department FROM students WHERE department IS NOT NULL UNION SELECT department FROM teachers WHERE department IS NOT NULL) departments ORDER BY department")
        departments = [row["department"] for row in cursor.fetchall() if row["department"]]
    return {"departments": departments, "locked": current_user.get("role") == "teacher"}

@app.get("/api/reports/export", dependencies=[Depends(require_role(["admin", "teacher"]))])
def export_attendance_report(
    department: Optional[str] = None,
    semester: Optional[str] = None,
    classCode: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: Optional[str] = "csv",
    current_user: dict = Depends(get_current_user),
):
    with db_session() as conn:
        cursor = conn.cursor()
        
        query = """
            SELECT
                r.rollnumber, r.studentname AS fullname, r.department, r.semester,
                s.classcode, s.classname, s.session_date AS date,
                a.timein, a.timeout, COALESCE(a.status, 'Absent') AS status,
                COALESCE(a.markedby, '-') AS markedby
            FROM attendance_sessions s
            JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            WHERE 1=1
        """
        params = []
        if current_user.get("role") == "teacher":
            query += " AND s.teacher_email = %s"
            params.append(current_user.get("email"))
        if department:
            query += " AND r.department = %s"
            params.append(department)
        if semester:
            query += " AND r.semester = %s"
            params.append(semester)
        if classCode:
            query += " AND s.classcode = %s"
            params.append(classCode)
        if start_date:
            query += " AND s.session_date >= %s"
            params.append(start_date)
        if end_date:
            query += " AND s.session_date <= %s"
            params.append(end_date)
            
        query += " ORDER BY s.session_date DESC, s.started_at DESC, r.rollnumber"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
    if format == "xlsx":
        import pandas as pd
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        data = []
        for r in rows:
            date_local = format_utc_to_local_date(r["timein"]) if r["timein"] not in ["-", "Pending", None] else r["date"]
            
            data.append({
                "Roll Number": r["rollnumber"],
                "Full Name": r["fullname"],
                "Department": r["department"],
                "Semester": f"Semester {r['semester']}" if r["semester"] else "-",
                "Class Code": r["classcode"],
                "Subject": r["classname"],
                "Date": date_local or "-",
                "Status": r["status"] or "Absent",
                "Marked By": r["markedby"] or "-"
            })
            
        df = pd.DataFrame(data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance Report')
            
            # Format the Excel sheet using openpyxl
            workbook = writer.book
            worksheet = writer.sheets['Attendance Report']
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(name='Roboto', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Deep blue
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            thin_side = Side(border_style="thin", color="D1D5DB")
            border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
                
            row_font = Font(name='Roboto', size=10)
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = row_font
                    cell.border = border
                    
                    val = str(cell.value or '')
                    if col_num in [1, 4, 5, 7, 8, 9]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
                    if col_num == 8: # Status
                        cell.alignment = center_align
                        if val == "Present":
                            cell.font = Font(name='Roboto', size=10, bold=True, color='15803D')
                            cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                        elif val == "Absent":
                            cell.font = Font(name='Roboto', size=10, bold=True, color='B91C1C')
                            cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                            
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        output.seek(0)
        filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    else:
        # Default to CSV
        import csv
        from io import StringIO
        from fastapi.responses import StreamingResponse
        
        output = StringIO()
        writer = csv.writer(output)
        
        writer.writerow([
            "Roll Number", "Full Name", "Department", "Semester",
            "Class Code", "Subject",
            "Date", "Status", "Marked By"
        ])
        
        for r in rows:
            date_local = format_utc_to_local_date(r["timein"]) if r["timein"] not in ["-", "Pending", None] else r["date"]
            
            writer.writerow([
                r["rollnumber"],
                r["fullname"],
                r["department"],
                f"Semester {r['semester']}" if r["semester"] else "-",
                r["classcode"],
                r["classname"],
                date_local or "-",
                r["status"] or "Absent",
                r["markedby"] or "-"
            ])
            
        output.seek(0)
        filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )



@app.get("/api/attendance/sessions", dependencies=[Depends(require_role(["admin", "teacher", "student"]))])
async def get_active_sessions(current_user: Optional[dict] = Depends(get_current_user_optional)):
    # Map displayed department names to class code prefixes
    dept_prefix_map = {
        'Computer Science': 'CS',
        'Information Technology': 'IT',
        'Electronics & Communication': 'EC',
        'Mechanical Engineering': 'ME',
        'Civil Engineering': 'CE',
        'Electrical Engineering': 'EE',
        'Humanities': 'HU'
    }

    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT active.classcode, active.classname, active.isactive,
                   history.location_required, history.allowed_radius_meters
            FROM active_sessions active
            LEFT JOIN LATERAL (
                SELECT location_required, allowed_radius_meters
                FROM attendance_sessions
                WHERE classcode = active.classcode AND isactive = TRUE
                ORDER BY id DESC LIMIT 1
            ) history ON TRUE
            ORDER BY active.classcode
        """)
        rows = cursor.fetchall()

    sessions = []
    for r in rows:
        sessions.append({
            "classCode": r["classcode"],
            "className": r["classname"],
            "isActive": r["isactive"],
            "locationRequired": bool(r.get("location_required")),
            "allowedRadiusMeters": r.get("allowed_radius_meters"),
        })

    # If no authenticated user info, return all
    if not current_user:
        return {"sessions": sessions}

    # Students should only see sessions matching their department + semester
    if current_user.get("role") == "student":
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT department, semester FROM students WHERE email = %s', (current_user.get("email"),))
            row = cursor.fetchone()
        if not row:
            return {"sessions": []}
        student_dept = row["department"]
        student_sem = normalize_semester(row.get("semester"))
        prefix = dept_prefix_map.get(student_dept) or department_prefix(student_dept)
        filtered = []
        for s in sessions:
            if not s["classCode"]: continue
            if prefix and s["classCode"].startswith(prefix + "-"):
                # extract semester digit from class code (e.g., CS-401 -> '4')
                try:
                    code_no = s["classCode"].split("-")[-1]
                    sess_sem = code_no[0]
                except Exception:
                    sess_sem = None
                if student_sem is None or sess_sem == student_sem:
                    filtered.append(s)
        return {"sessions": filtered}

    # Teachers should only see sessions for their department
    if current_user.get("role") == "teacher":
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT department, subjects FROM teachers WHERE email = %s', (current_user.get("email"),))
            row = cursor.fetchone()
        if not row:
            return {"sessions": []}
        teacher_dept = row["department"]
        prefix = dept_prefix_map.get(teacher_dept) or department_prefix(teacher_dept)
        filtered = []
        for s in sessions:
            if not s["classCode"]: continue
            if prefix and s["classCode"].startswith(prefix + "-"):
                filtered.append(s)
        return {"sessions": filtered}

    return {"sessions": sessions}

@app.post("/api/attendance/sessions/toggle", dependencies=[Depends(require_role(["admin", "teacher"]))])
async def toggle_active_session(payload: ToggleSessionRequest, current_user: dict = Depends(get_current_user)):
    # Map department names to class code prefixes
    dept_prefix_map = {
        'Computer Science': 'CS',
        'Information Technology': 'IT',
        'Electronics & Communication': 'EC',
        'Mechanical Engineering': 'ME',
        'Civil Engineering': 'CE',
        'Electrical Engineering': 'EE',
        'Humanities': 'HU'
    }

    # Admin can toggle any session
    if current_user.get("role") != "admin":
        # teacher - verify department match
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT department, fullname FROM teachers WHERE email = %s', (current_user.get("email"),))
            teacher_row = cursor.fetchone()
        if not teacher_row:
            raise HTTPException(status_code=403, detail="Teacher profile not found")
        teacher_dept = teacher_row["department"]
        prefix = dept_prefix_map.get(teacher_dept) or department_prefix(teacher_dept)
        if not prefix:
            raise HTTPException(status_code=403, detail="Teacher department is not configured for attendance sessions")
        if not payload.classCode.startswith(prefix + "-"):
            raise HTTPException(status_code=403, detail="You can only toggle sessions for your own department")

    if payload.active and payload.locationRequired:
        if payload.latitude is None or payload.longitude is None:
            raise HTTPException(status_code=400, detail="Teacher location is required to open this attendance window.")
        if not (-90 <= payload.latitude <= 90 and -180 <= payload.longitude <= 180):
            raise HTTPException(status_code=400, detail="Invalid teacher location coordinates.")
        if not payload.allowedRadiusMeters or not (20 <= payload.allowedRadiusMeters <= 1000):
            raise HTTPException(status_code=400, detail="Allowed radius must be between 20 and 1000 metres.")

    teacher_dept = teacher_row["department"] if current_user.get("role") != "admin" else payload.classCode.split("-")[0]
    teacher_name = teacher_row["fullname"] if current_user.get("role") != "admin" else current_user.get("fullName", "Administrator")
    session_semester = payload.classCode.split("-")[-1][:1]
    session_id = None

    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT classname FROM active_sessions WHERE classcode = %s", (payload.classCode,))
        class_row = cursor.fetchone()
        if not class_row:
            raise HTTPException(status_code=404, detail="Attendance session not found")

        # A teacher can run only one live attendance window at a time. Close
        # any older window before activating the newly selected class.
        if payload.active:
            cursor.execute("""
                UPDATE attendance_sessions
                SET isactive = FALSE, ended_at = CURRENT_TIMESTAMP
                WHERE isactive = TRUE
                  AND classcode <> %s
                  AND (
                      teacher_email = %s
                      OR (LOWER(department) = LOWER(%s) AND semester = %s)
                  )
                RETURNING classcode
            """, (
                payload.classCode,
                current_user.get("email"),
                teacher_dept,
                session_semester,
            ))
            closed_classcodes = {row["classcode"] for row in cursor.fetchall()}
            for closed_classcode in closed_classcodes:
                cursor.execute("""
                    UPDATE active_sessions
                    SET isactive = FALSE
                    WHERE classcode = %s
                      AND NOT EXISTS (
                          SELECT 1 FROM attendance_sessions
                          WHERE classcode = %s AND isactive = TRUE
                      )
                """, (closed_classcode, closed_classcode))

        cursor.execute("""
            UPDATE active_sessions
            SET isactive = %s
            WHERE classcode = %s
        """, (payload.active, payload.classCode))

        if payload.active:
            cursor.execute("""
                SELECT id FROM attendance_sessions
                WHERE classcode = %s AND teacher_email = %s AND isactive = TRUE
                ORDER BY id DESC LIMIT 1
            """, (payload.classCode, current_user.get("email")))
            existing = cursor.fetchone()
            if existing:
                session_id = existing["id"]
            else:
                cursor.execute("""
                    INSERT INTO attendance_sessions (
                        classcode, classname, department, semester,
                        teacher_email, teacher_name, session_date, isactive,
                        latitude, longitude, allowed_radius_meters, location_required
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    payload.classCode,
                    class_row["classname"],
                    teacher_dept,
                    session_semester,
                    current_user.get("email"),
                    teacher_name,
                    datetime.now(LOCAL_TZ).strftime("%Y-%m-%d"),
                    payload.latitude,
                    payload.longitude,
                    payload.allowedRadiusMeters or 100,
                    bool(payload.locationRequired),
                ))
                session_id = cursor.fetchone()["id"]
                cursor.execute("""
                    INSERT INTO attendance_session_roster (
                        sessionid, studentid, rollnumber, studentname,
                        department, semester
                    )
                    SELECT %s, id, rollnumber, fullname, department, semester::text
                    FROM students
                    WHERE status = 'Active'
                      AND LOWER(department) = LOWER(%s)
                      AND semester::text = %s
                    ON CONFLICT (sessionid, rollnumber) DO NOTHING
                """, (session_id, teacher_dept, session_semester))
        else:
            cursor.execute("""
                UPDATE attendance_sessions
                SET isactive = FALSE, ended_at = CURRENT_TIMESTAMP
                WHERE classcode = %s AND teacher_email = %s AND isactive = TRUE
            """, (payload.classCode, current_user.get("email")))
        conn.commit()
    print(f"[SessionToggle] User={current_user.get('email')} set {payload.classCode} active={payload.active}")
    return {
        "success": True,
        "classCode": payload.classCode,
        "isActive": payload.active,
        "sessionId": session_id,
    }

# ─── Live Webcam Scan API ─────────────────────────────────────
@app.post("/api/attendance/scan")
async def scan_attendance(request: ScanRequest, current_user: Optional[dict] = Depends(get_current_user_optional)):
    session_context = None
    # Phase 1: Quick DB check — is the session active? Release connection immediately.
    if request.classCode:
        with db_session() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT isactive FROM active_sessions WHERE classcode = %s", (request.classCode,))
            row = cursor.fetchone()
            if not row or not row["isactive"]:
                raise HTTPException(status_code=400, detail="Attendance window is closed for this class.")
            cursor.execute("""
                SELECT id, department, semester, latitude, longitude,
                       allowed_radius_meters, location_required
                FROM attendance_sessions
                WHERE classcode = %s AND isactive = TRUE
                ORDER BY id DESC LIMIT 1
            """, (request.classCode,))
            session_context = cursor.fetchone()
            if not session_context:
                raise HTTPException(
                    status_code=400,
                    detail="Class session record is unavailable. Reopen the attendance window."
                )

            if session_context["location_required"]:
                if request.latitude is None or request.longitude is None or request.locationAccuracy is None:
                    raise HTTPException(status_code=400, detail="Location permission is required for this attendance session.")
                if not (-90 <= request.latitude <= 90 and -180 <= request.longitude <= 180):
                    raise HTTPException(status_code=400, detail="Invalid location coordinates.")
                max_accuracy = float(os.environ.get("MAX_LOCATION_ACCURACY_METERS", "100"))
                distance_meters = haversine_distance_meters(
                    session_context["latitude"], session_context["longitude"],
                    request.latitude, request.longitude,
                )
                location_reason = None
                if request.locationAccuracy <= 0 or request.locationAccuracy > max_accuracy:
                    location_reason = f"Location accuracy is too low ({request.locationAccuracy:.0f}m). Move near a window and retry."
                elif distance_meters > session_context["allowed_radius_meters"]:
                    location_reason = (
                        f"You are {distance_meters:.0f}m from the classroom; "
                        f"the allowed radius is {session_context['allowed_radius_meters']:.0f}m."
                    )
                if location_reason:
                    cursor.execute("""
                        INSERT INTO attendance_verification_audit (
                            sessionid, user_email, classcode, outcome, reason,
                            submitted_latitude, submitted_longitude, location_accuracy, distance_meters
                        ) VALUES (%s, %s, %s, 'rejected', %s, %s, %s, %s, %s)
                    """, (
                        session_context["id"], current_user.get("email") if current_user else None,
                        request.classCode, location_reason, request.latitude, request.longitude,
                        request.locationAccuracy, distance_meters,
                    ))
                    conn.commit()
                    raise HTTPException(status_code=403, detail=location_reason)
                session_context["distance_meters"] = distance_meters

    # Phase 2: CPU-heavy OpenCV inference — NO DB connection held
    frame = decode_base64_image(request.image)
    if frame is None:
        return {"faces": []}

    embedding, box, frame_quality_live, liveness_metrics = await run_in_threadpool(extract_face_embedding, frame)
    
    updated_faces = []
    
    if not embedding:
        return {"faces": updated_faces}

    # Phase 3: Quick DB match + insert — short-lived connection
    with db_session() as conn:
        cursor = conn.cursor()
        
        # Consistent Timezone Handling: Determine today's date in local time
        today_local = datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")
        utc_start, utc_end = get_utc_range_for_local_date(today_local)
        
        match_params = [str(embedding)]
        match_filters = ["embedding IS NOT NULL", "status = 'Active'"]
        if session_context:
            match_filters.extend([
                "LOWER(department) = LOWER(%s)",
                "semester::text = %s",
            ])
            match_params.extend([
                session_context["department"],
                str(session_context["semester"]),
            ])
        # Self-attendance already has an authenticated identity. Compare only
        # against that student's enrolled template instead of allowing another
        # roster member to become the nearest candidate on a noisy frame.
        if current_user and current_user.get("role") == "student":
            match_filters.append("rollnumber = %s")
            match_params.append(current_user.get("referenceId"))
            if session_context:
                match_filters.append("EXISTS (SELECT 1 FROM attendance_session_roster eligible WHERE eligible.sessionid = %s AND eligible.studentid = students.id)")
                match_params.append(session_context["id"])
        cursor.execute(f"""
            SELECT rollnumber, fullname, department, semester,
                   (embedding <=> %s::vector) AS distance
            FROM students
            WHERE {' AND '.join(match_filters)}
            ORDER BY distance ASC LIMIT 1;
        """, tuple(match_params))
        match = cursor.fetchone()
        
        face_info = {
            "box": box,
            "name": "Unknown",
            "is_live": False,
            "liveness_status": "verifying",
            "liveness_prompt": "Hold still while your face is verified",
            "identity_verified": True,
            "marked": False,
            "distance": 1.0
        }

        # Configurable Biometrics: Move threshold to env variable
        match_threshold = (
            SELF_FACE_MATCH_THRESHOLD
            if current_user and current_user.get("role") == "student"
            else FACE_MATCH_THRESHOLD
        )
        if match and match["distance"] < match_threshold:
            roll = match["rollnumber"]
            friendly_name = match["fullname"]
            dept = match["department"]
            
            face_info["name"] = friendly_name
            face_info["distance"] = match["distance"]

            challenge_key = f"{current_user.get('email') if current_user else 'anonymous'}:{request.classCode or 'none'}:{roll}"
            is_live, liveness_status, liveness_prompt = verify_motion_liveness(
                challenge_key,
                liveness_metrics.get("yaw_ratio", 1.0),
                frame_quality_live,
            )
            face_info["is_live"] = is_live
            face_info["liveness_status"] = liveness_status
            face_info["liveness_prompt"] = liveness_prompt

            if is_live:
                if current_user and current_user.get("role") == "student":
                    expected_roll = current_user.get("referenceId")
                    expected_name = current_user.get("fullName")
                    if roll != expected_roll and friendly_name != expected_name:
                        print(f"Identity mismatch in scan: expected {expected_roll}, got {roll}")
                        face_info["identity_verified"] = False
                        face_info["is_live"] = False
                        face_info["name"] = f"Mismatch: {friendly_name}"
                        updated_faces.append(face_info)
                        return {"faces": updated_faces}
                
                # If a teacher performed the scan, record their full name; otherwise label as Webcam
                if current_user and current_user.get("role") == "teacher":
                    marked_by_str = current_user.get("fullName") or f"Teacher ({current_user.get('referenceId')})"
                else:
                    marked_by_str = f"Webcam ({request.classCode})" if request.classCode else "Webcam"
                
                session_id = session_context["id"] if session_context else None
                cursor.execute("""
                    SELECT COUNT(*) FROM attendance
                    WHERE rollnumber = %s
                      AND sessionid = %s
                      AND status = 'Present'
                """, (roll, session_id))
                
                if cursor.fetchone()[0] == 0:
                    # Force UTC timestamp internally
                    utc_now_iso = get_utc_now().isoformat()
                    cursor.execute("""
                        INSERT INTO attendance (
                            rollnumber, studentname, department, date,
                            timein, timeout, status, markedby, sessionid,
                            submitted_latitude, submitted_longitude, location_accuracy,
                            distance_meters, face_confidence, verification_method
                        )
                        VALUES (%s, %s, %s, %s, %s, 'Pending', 'Present', %s, %s,
                                %s, %s, %s, %s, %s, 'face+liveness+geofence')
                        ON CONFLICT(sessionid, rollnumber) WHERE sessionid IS NOT NULL
                        DO UPDATE SET status='Present', timein=%s, markedby=%s,
                            submitted_latitude=%s, submitted_longitude=%s,
                            location_accuracy=%s, distance_meters=%s,
                            face_confidence=%s, verification_method='face+liveness+geofence'
                    """, (
                        roll, friendly_name, dept, today_local, utc_now_iso,
                        marked_by_str, session_id, request.latitude, request.longitude,
                        request.locationAccuracy, session_context.get("distance_meters") if session_context else None,
                        max(0.0, 1.0 - float(match["distance"])),
                        utc_now_iso, marked_by_str, request.latitude, request.longitude,
                        request.locationAccuracy, session_context.get("distance_meters") if session_context else None,
                        max(0.0, 1.0 - float(match["distance"])),
                    ))
                    conn.commit()
                    face_info["marked"] = True
                
        updated_faces.append(face_info)
        
    return {"faces": updated_faces}


# ─── Reports Analytics Summary ─────────────────────────────────
@app.get("/api/reports/summary", dependencies=[Depends(require_role(["admin"]))])
async def get_reports_summary():
    with db_session() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                s.department,
                COUNT(DISTINCT r.studentid) AS students,
                COUNT(r.rollnumber) AS roster_entries,
                COUNT(a.rollnumber) FILTER (WHERE a.status = 'Present') AS present
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            GROUP BY s.department
            ORDER BY s.department
        """)
        department_stats = []
        for row in cursor.fetchall():
            roster_entries = int(row["roster_entries"] or 0)
            present = int(row["present"] or 0)
            absent = max(roster_entries - present, 0)
            department_stats.append({
                "name": row["department"],
                "totalStudents": int(row["students"] or 0),
                "present": present,
                "absent": absent,
                "percentage": round((present / roster_entries) * 100, 1) if roster_entries else 0.0,
            })

        cursor.execute("SELECT COUNT(*) FROM students WHERE status = 'Active'")
        total_students = cursor.fetchone()[0]

        cursor.execute("""
            SELECT
                COUNT(r.rollnumber) AS total,
                COUNT(a.rollnumber) FILTER (WHERE a.status = 'Present') AS present
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
        """)
        totals = cursor.fetchone()
        total_records = int(totals["total"] or 0)
        total_present = int(totals["present"] or 0)
        total_absent = max(total_records - total_present, 0)
        avg_attendance = round((total_present / total_records) * 100, 1) if total_records else 0.0

        cursor.execute("""
            SELECT
                s.session_date AS date,
                COUNT(r.rollnumber) AS total,
                COUNT(a.rollnumber) FILTER (WHERE a.status = 'Present') AS present
            FROM attendance_sessions s
            LEFT JOIN attendance_session_roster r ON r.sessionid = s.id
            LEFT JOIN attendance a
              ON a.sessionid = s.id AND a.rollnumber = r.rollnumber
            GROUP BY s.session_date
            ORDER BY s.session_date DESC
            LIMIT 7
        """)
        overview = []
        for row in reversed(cursor.fetchall()):
            date_label = datetime.strptime(row["date"], "%Y-%m-%d").strftime("%b %d")
            rate = round((row["present"] / row["total"]) * 100, 1) if row["total"] else 0.0
            overview.append({"date": date_label, "value": rate})
            
    return {
        "totalStudents": total_students,
        "averageAttendance": avg_attendance,
        "totalPresent": total_present,
        "totalAbsent": total_absent,
        "overview": overview,
        "departmentStats": department_stats
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
